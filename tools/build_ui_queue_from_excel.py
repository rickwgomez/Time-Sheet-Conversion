import argparse
import csv
import json
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook


SKIP_JOB_VALUES = {"JOB# NEEDED", "", None}
SKIP_EMPLOYEE_VALUES = {"LEAD NAME NEEDED", "", None}
SKIP_TIME_VALUES = {"ON SITE MISSING", "", None}
DEFAULT_RULES = Path("config") / "reading_rules.json"


def parse_section_header(value):
    if not isinstance(value, str) or " / " not in value:
        return None
    date_text, work_performed = value.split(" / ", 1)
    return date_text.strip(), work_performed.strip()


def load_employee_directory(rules_path):
    with rules_path.open("r", encoding="utf-8") as handle:
        rules = json.load(handle)
    directory = {}
    for employee in rules.get("employee_names", []):
        first_name = employee.get("canonical", "").strip()
        last_name = employee.get("last_name", "").strip()
        if not first_name:
            continue
        full_name = f"{first_name} {last_name}".strip()
        employee_fields = {
            "employee_name": full_name,
        }
        directory[first_name.lower()] = employee_fields
        directory[full_name.lower()] = employee_fields
        for alias in employee.get("aliases", []):
            directory[str(alias).strip().lower()] = employee_fields
    return directory


def is_monday_job_number(job_number, date_text, rules):
    job_rules = rules.get("job_number_rules", {})
    expected_digits = int(job_rules.get("expected_digits", 7))
    job_text = str(job_number)
    if not job_text.isdigit() or len(job_text) != expected_digits:
        return False
    if not isinstance(date_text, str) or len(date_text) < 4 or not date_text[:4].isdigit():
        return False
    work_year = int(date_text[:4])
    allowed_prefixes = {str(work_year)[2:4], str(work_year - 1)[2:4]}
    return job_text[:2] in allowed_prefixes


def expand_employee_name(employee_name, directory):
    first_name = str(employee_name).strip()
    match = directory.get(first_name.lower())
    if match:
        return match
    parts = first_name.split()
    if len(parts) >= 2:
        return {
            "employee_name": first_name,
        }
    return {
        "employee_name": first_name,
    }


def build_queue(workbook_path, rules_path=DEFAULT_RULES):
    with rules_path.open("r", encoding="utf-8") as handle:
        rules = json.load(handle)
    employee_directory = load_employee_directory(rules_path)
    temp_dir = Path(tempfile.mkdtemp(prefix="timecard_excel_"))
    temp_workbook = temp_dir / workbook_path.name
    shutil.copy2(workbook_path, temp_workbook)
    wb = load_workbook(temp_workbook, data_only=True)
    ws = wb["Weekly On Site"]

    current_date = None
    current_work_performed = None
    entries = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        first_cell = row[0]
        section = parse_section_header(first_cell)
        if section:
            current_date, current_work_performed = section
            continue

        if first_cell in (None, "Job#", "Weekly On Site Time") or current_date is None:
            continue

        job_number, employee_name, start_time, end_time, total_hours = row[:5]
        if job_number in SKIP_JOB_VALUES or not str(job_number).isdigit():
            continue
        if not is_monday_job_number(job_number, current_date, rules):
            continue
        if employee_name in SKIP_EMPLOYEE_VALUES:
            continue
        if start_time in SKIP_TIME_VALUES or end_time in SKIP_TIME_VALUES:
            continue
        if total_hours in ("", None):
            continue

        employee_fields = expand_employee_name(employee_name, employee_directory)
        entries.append({
            **employee_fields,
            "Project": str(job_number),
            "Hours": float(total_hours),
            "Date": current_date,
            "Type": "Billable",
            "Description": current_work_performed,
        })

    return {"source_workbook": str(workbook_path), "source_sheet": "Weekly On Site", "entries": entries}


def write_csv(path, entries):
    fieldnames = [
        "Employee",
        "Project",
        "Hours",
        "Date",
        "Type",
        "Description",
    ]
    rows = [
        {
            "Employee": entry["employee_name"],
            "Project": entry["Project"],
            "Hours": entry["Hours"],
            "Date": entry["Date"],
            "Type": entry["Type"],
            "Description": entry["Description"],
        }
        for entry in entries
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Build Vibe/Monday UI entry queue from Weekly On Site Excel tab.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--rules", type=Path, default=DEFAULT_RULES)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--csv-output", type=Path)
    args = parser.parse_args()

    output = args.output
    if output is None:
        output = Path("Monday Output") / f"{args.workbook.stem}.ui_entry_queue.json"

    csv_output = args.csv_output
    if csv_output is None:
        csv_output = Path("Monday Output") / f"{args.workbook.stem}.upload.csv"

    payload = build_queue(args.workbook, args.rules)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
        handle.write("\n")
    csv_output.parent.mkdir(parents=True, exist_ok=True)
    write_csv(csv_output, payload["entries"])
    print(f"Wrote {output} with {len(payload['entries'])} entries.")
    print(f"Wrote {csv_output} with {len(payload['entries'])} entries.")


if __name__ == "__main__":
    main()
